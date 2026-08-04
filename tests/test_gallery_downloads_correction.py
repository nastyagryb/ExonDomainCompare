from __future__ import annotations

import io
import json
import sys
import tomllib
import zipfile
from pathlib import Path

import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))
sys.path.insert(0, str(ROOT / "scripts" / "shared_gene_analysis"))

SINGLE = ROOT / "runs" / "2026-07-23_1100_fgfr1_gallus_core_pilot"
MULTI = ROOT / "runs" / "2026-07-26_2157_fgfr1_gallus_mus_core_pilot"
TP53 = ROOT / "runs" / "2026-07-21_1436_custom_run"

GALLUS = "gallus_gallus"
MUS = "mus_musculus"


def _package_path(job) -> Path:
    from exondomaincompare.shared_gene_analysis.package_builder import PACKAGES_ROOT
    assert job.zip_path.startswith("package:")
    return PACKAGES_ROOT / job.zip_path[len("package:"):]

SINGLE_CATEGORIES = {
    "Exon structure", "Isoform analysis", "Domain architecture",
    "Exon–domain boundaries", "Genomic context", "Exploratory candidates",
}

# The final comparative inventory (Part 5).
COMPARATIVE_MAIN = {
    "cmp_msa_aligned_exon_architecture",
    "cmp_primary_msa_overview",
    "cmp_domain_architecture_msa",
    "cmp_domain_architecture_native",
    "cmp_exon_domain_architecture_native",
    "cmp_exon_domain_architecture_msa",
    "cmp_boundary_matrix",
    "cmp_paired_signed_distance",
    "cmp_boundary_position_consistency",
    "cmp_comparative_synteny",
    "cmp_isoform_diversity",
}
COMPARATIVE_SUPPLEMENTS = {
    "cmp_native_exon_architecture",
    "cmp_domain_annotation_matrix",
    "cmp_local_boundary_architecture",
    "cmp_synteny_neighbour_conservation",
}
# Text-only pages that must never come back as permanent cards.
RETIRED_CARDS = {"cmp_pairwise_identity", "cmp_exon_boundary_alignment_summary"}


def _figures(run_dir: Path) -> list:
    idx = run_dir / "website_indices" / "figures_index.json"
    if not idx.is_file():
        pytest.skip(f"no figure index for {run_dir.name}")
    return json.loads(idx.read_text()).get("figures") or []


def _schema(run_dir: Path, species_id: str) -> list:
    return sorted(
        (f.get("category"), f.get("kind"), f.get("figure_type") or f.get("figure_id"))
        for f in _figures(run_dir)
        if (f.get("scope") or "") != "comparative"
        and (f.get("species_id") or "") == species_id
    )


# --------------------------------------------------------------------------- #
# Part 2 — species Scope reuses the accepted single-species pipeline
# --------------------------------------------------------------------------- #

def test_gallus_scope_schema_matches_standalone_gallus_gallery():
    if not SINGLE.is_dir() or not MULTI.is_dir():
        pytest.skip("both FGFR1 reference runs are required")
    standalone = _schema(SINGLE, GALLUS)
    in_multi = _schema(MULTI, GALLUS)
    assert standalone, "the accepted standalone Gallus Gallery is empty"
    assert in_multi == standalone


def test_mus_scope_uses_the_same_logic_on_real_mus_data():
    if not SINGLE.is_dir() or not MULTI.is_dir():
        pytest.skip("both FGFR1 reference runs are required")
    assert _schema(MULTI, MUS) == _schema(SINGLE, GALLUS)
    # Real Mus data, not copied Gallus files.
    mus_cards = [f for f in _figures(MULTI) if f.get("species_id") == MUS]
    gallus_ids = {f.get("figure_id") for f in _figures(MULTI)
                  if f.get("species_id") == GALLUS}
    assert mus_cards
    for card in mus_cards:
        assert card["figure_id"] not in gallus_ids
        assert MUS in card["figure_id"]
        assert card.get("protein_id")
        assert card.get("protein_id") not in {"", None}
    mus_proteins = {c.get("protein_id") for c in mus_cards}
    gallus_proteins = {c.get("protein_id") for c in _figures(MULTI)
                       if c.get("species_id") == GALLUS}
    assert mus_proteins.isdisjoint(gallus_proteins)


def test_species_scope_keeps_the_six_single_species_categories():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    for species in (GALLUS, MUS):
        cats = {c for c, _kind, _type in _schema(MULTI, species)}
        assert cats == SINGLE_CATEGORIES, species


def test_species_scope_uses_the_same_renderer_and_keeps_one_supplement():
    if not SINGLE.is_dir() or not MULTI.is_dir():
        pytest.skip("both FGFR1 reference runs are required")

    def cards_by_type(run: Path, species: str) -> dict:
        return {f.get("figure_type") or f["figure_id"]: f for f in _figures(run)
                if f.get("species_id") == species
                and (f.get("scope") or "") != "comparative"}

    reference = cards_by_type(SINGLE, GALLUS)
    for run, species in ((SINGLE, GALLUS), (MULTI, GALLUS), (MULTI, MUS)):
        cards = cards_by_type(run, species)
        assert len(cards) == 15, (run.name, species, len(cards))
        supplements = [c for c in cards.values() if c.get("kind") == "supplement"]
        assert len(supplements) == 1, (run.name, species)
        for figure_type, card in cards.items():
            # A figure type must be drawn by the same renderer everywhere, so a
            # species Scope cannot silently fall back to a different drawing path.
            assert card.get("renderer") == reference[figure_type].get("renderer"), (
                run.name, species, figure_type)
            # Vector-safe SVG/PDF plus a raster preview for every card.
            assert card.get("svg_url") and card.get("pdf_url") and card.get("png_url")


def test_isoform_analysis_has_three_cards_per_species():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    for species in (GALLUS, MUS):
        iso = [f for f in _figures(MULTI)
               if f.get("species_id") == species
               and f.get("category") == "Isoform analysis"]
        assert len(iso) == 3, (species, [c["figure_id"] for c in iso])


def test_one_shared_figure_sequence_builds_every_gallery():
    from plotting.figure_sequence import FIGURE_STAGES
    names = [module for _label, module in FIGURE_STAGES]
    assert names[0] == "plotting.generate_shared_main_figures"
    # The comparative stage is the last producer; registration follows it and decides
    # which of the produced cards a reader may actually see.
    assert names[-2] == "plotting.generate_comparative_gallery_figures"
    assert names[-1] == "plotting.figure_registration"
    # The pipeline must not keep its own private copy of the stage list.
    pipeline = (ROOT / "src" / "exondomaincompare" / "framework"
                / "run_core_gene_analysis.py").read_text()
    assert "run_figure_stages" in pipeline
    assert "plotting.generate_exon_map_figures" not in pipeline


# --------------------------------------------------------------------------- #
# Parts 3–5 — comparative inventory contains no text-only placeholder
# --------------------------------------------------------------------------- #

def _comparative_cards() -> list:
    return [f for f in _figures(MULTI) if (f.get("scope") or "") == "comparative"]


def test_final_comparative_inventory():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    cards = _comparative_cards()
    ids = {c["figure_id"] for c in cards}
    assert ids == COMPARATIVE_MAIN | COMPARATIVE_SUPPLEMENTS
    mains = {c["figure_id"] for c in cards if c.get("kind") == "main"}
    assert mains == COMPARATIVE_MAIN


def test_text_only_comparative_cards_are_gone():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    ids = {c["figure_id"] for c in _comparative_cards()}
    assert ids.isdisjoint(RETIRED_CARDS)
    fig_dir = MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
    for retired in RETIRED_CARDS:
        assert not list(fig_dir.glob(f"{retired}.*")), retired


def test_comparative_synteny_and_isoform_plots_exist_and_are_plots():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    fig_dir = MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
    for stem, floor in (("cmp_comparative_synteny", 4000),
                        ("cmp_isoform_diversity", 3000),
                        ("cmp_primary_msa_overview", 20000)):
        svg = fig_dir / f"{stem}.svg"
        assert svg.is_file(), stem
        text = svg.read_text(encoding="utf-8")
        assert svg.stat().st_size > floor, (stem, svg.stat().st_size)
        # A real plot draws geometry; a text page would be <text> only.
        assert text.count("<rect") + text.count("<line") + text.count("<circle") > 12, stem


def test_comparative_domain_figures_carry_readable_labels():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    fig_dir = MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
    for stem in ("cmp_domain_architecture_msa", "cmp_domain_architecture_native"):
        text = (fig_dir / f"{stem}.svg").read_text(encoding="utf-8")
        # Domain names, not anonymous rectangles.
        assert "Ig-like domain 1" in text, stem
        assert "kinase" in text.lower(), stem
        # Species identity and the protein accession are on the figure.
        assert "Gallus gallus" in text and "Mus musculus" in text, stem


def test_msa_overview_reports_identity_as_a_metric():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    svg = (MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
           / "cmp_primary_msa_overview.svg").read_text(encoding="utf-8")
    assert "Pairwise identity" in svg
    assert "mismatch" in svg and "indel" in svg


def test_boundary_matrix_matches_the_explorer_values():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    tsv = (MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
           / "cmp_boundary_matrix.tsv")
    if not tsv.is_file():
        pytest.skip("matrix source table not rendered")
    # The matrix table is wide: one row per species, one column per comparable
    # group, exactly like the cells of the interactive Explorer.
    lines = tsv.read_text().splitlines()
    header = lines[0].split("\t")
    groups = header[2:]
    assert groups and all(g.startswith("CBG") for g in groups), header
    from exondomaincompare.shared_gene_analysis import boundary_observations as bo
    canonical = {
        (r["species_id"], r["comparable_boundary_group_id"]): str(r["signed_distance"])
        for r in bo.build_rows(MULTI)
        if str(r["comparable_boundary_group_id"]).startswith("CBG")
    }
    checked = 0
    for line in lines[1:]:
        cells = line.split("\t")
        species = cells[0]
        for group, value in zip(groups, cells[2:]):
            if value in ("", "—", "unavailable"):
                continue
            assert value == canonical[(species, group)], (species, group)
            checked += 1
    assert checked >= len(groups)


def test_paired_signed_distance_shows_both_species():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    svg = (MULTI / "results" / "generic_gene_analysis" / "figures" / "comparative"
           / "cmp_paired_signed_distance.svg").read_text(encoding="utf-8")
    assert "Gallus gallus" in svg and "Mus musculus" in svg
    # One marker per species per group, not a single difference bar.
    assert svg.count("<circle") >= 8


# --------------------------------------------------------------------------- #
# Part 8 — Boundary observation tables
# --------------------------------------------------------------------------- #

REQUIRED_BOUNDARY_FIELDS = [
    "species_id", "scientific_name", "primary_protein", "boundary_id",
    "comparable_boundary_group_id", "exon_transition", "native_aa_position",
    "MSA_column", "nearest_domain_instance_id", "nearest_domain_label",
    "nearest_edge", "signed_distance", "absolute_distance", "boundary_class",
    "mapping_method", "mapping_confidence", "status",
]


def test_boundary_observation_columns_are_the_download_contract():
    from exondomaincompare.shared_gene_analysis import boundary_observations as bo
    assert list(bo.COLUMNS) == REQUIRED_BOUNDARY_FIELDS


def test_boundary_observations_exist_for_a_single_species_run():
    if not SINGLE.is_dir():
        pytest.skip("standalone Gallus run required")
    from exondomaincompare.shared_gene_analysis import boundary_observations as bo
    rows = bo.build_rows(SINGLE)
    assert rows
    assert {r["species_id"] for r in rows} == {GALLUS}
    assert all(r["comparable_boundary_group_id"] == "not_applicable_single_species"
               for r in rows)
    out = bo.ensure_table(SINGLE, GALLUS)
    assert out and out.is_file()
    assert out.read_text().splitlines()[0].split("\t") == REQUIRED_BOUNDARY_FIELDS
    assert bo.label_for(False) == "Boundary observations (TSV)"


def test_multi_species_boundary_long_table_has_one_row_per_observation():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis import boundary_observations as bo
    rows = bo.build_rows(MULTI)
    assert {r["species_id"] for r in rows} == {GALLUS, MUS}
    keys = [(r["species_id"], r["boundary_id"]) for r in rows]
    assert len(keys) == len(set(keys))
    mapped = [r for r in rows
              if r["comparable_boundary_group_id"] not in ("unmapped",
                                                           "not_applicable_single_species")]
    assert mapped, "no observation carries a comparable-boundary group"
    for row in mapped:
        assert row["MSA_column"] != "unavailable"
        assert row["mapping_method"] not in ("", "unmapped")
    assert bo.label_for(True) == "All species Boundary observations (TSV)"


# --------------------------------------------------------------------------- #
# Parts 6 / 7 — scope correction and the availability contract
# --------------------------------------------------------------------------- #

COMPARATIVE_ONLY_ITEMS = {
    "species_inventory", "primary_proteins_msa", "msa_aligned_exons",
    "comparable_domain_groups", "boundary_consistency", "comparative_synteny",
    "excel_workbook", "comparative_figures",
}


def test_single_species_run_has_no_comparative_scope_or_items():
    if not SINGLE.is_dir():
        pytest.skip("standalone Gallus run required")
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities, scopes_for_run
    scopes = {s["id"] for s in scopes_for_run(SINGLE)}
    assert scopes == {GALLUS}
    caps = capabilities(SINGLE)
    assert caps["multi_species"] is False
    assert caps["scope"] == GALLUS
    assert set(caps["items"]).isdisjoint(COMPARATIVE_ONLY_ITEMS)
    # Nothing on the page may be offered without being deliverable.
    for item in caps["items"].values():
        assert item["available"] or item["reason"]


def test_single_species_boundary_observations_are_available():
    if not SINGLE.is_dir():
        pytest.skip("standalone Gallus run required")
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities
    item = capabilities(SINGLE)["items"]["boundary_observations"]
    assert item["available"] is True
    assert item["label"] == "Boundary observations (TSV)"
    assert item["path"] and (ROOT / item["path"]).is_file()


def test_multi_species_boundary_long_table_is_available():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities
    item = capabilities(MULTI, "comparative")["items"]["boundary_long_table"]
    assert item["available"] is True
    assert item["label"] == "All species Boundary observations (TSV)"


def test_multi_species_scopes_and_species_scope_files():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities, scopes_for_run
    ids = [s["id"] for s in scopes_for_run(MULTI)]
    assert ids == ["comparative", "all", GALLUS, MUS]
    # A species Scope offers the same item set as a single-species dataset.
    species_items = set(capabilities(MULTI, GALLUS)["items"])
    if SINGLE.is_dir():
        assert species_items == set(capabilities(SINGLE)["items"])
    assert species_items.isdisjoint(COMPARATIVE_ONLY_ITEMS)
    # The comparative Scope offers only comparative products.
    comparative_items = set(capabilities(MULTI, "comparative")["items"])
    assert "isoform_alignment" not in comparative_items
    assert "boundary_long_table" in comparative_items


def test_unavailable_items_are_excluded_from_presets_and_resolution(tmp_path):
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities
    run = tmp_path / "2026-01-01_0000_empty"
    (run / "website_indices" / "generic").mkdir(parents=True)
    (run / "website_indices" / "generic" / "protein_coordinate_model.json").write_text(
        json.dumps({"gene_symbol": "TESTG",
                    "models": [{"species_id": "test_species",
                                "scientific_name": "Test species",
                                "protein_id": "NP_1"}]}))
    caps = capabilities(run)
    assert caps["scope"] == "test_species"
    unavailable = [i for i, v in caps["items"].items() if not v["available"]]
    assert unavailable, "this fixture is meant to have unavailable items"
    for iid in unavailable:
        assert caps["items"][iid]["reason"]
        assert caps["items"][iid]["path"] is None
        for preset in caps["presets"].values():
            assert iid not in preset["items"]
            if iid in (preset.get("unavailable_items") or []):
                assert iid not in preset["items"]


def test_build_drops_unavailable_items_with_a_reason(tmp_path):
    from exondomaincompare.shared_gene_analysis.package_builder import build_package
    run = tmp_path / "2026-01-01_0000_empty"
    (run / "website_indices" / "generic").mkdir(parents=True)
    (run / "website_indices" / "generic" / "protein_coordinate_model.json").write_text(
        json.dumps({"gene_symbol": "TESTG",
                    "models": [{"species_id": "test_species",
                                "protein_id": "NP_1"}]}))
    job = build_package(run, {"scope": "test_species", "preset": "custom",
                              "items": ["isoform_alignment"]})
    assert job.status == "failed"
    assert "available" in job.error
    assert all(i not in job.selected_items for i in ["isoform_alignment"])


# --------------------------------------------------------------------------- #
# Parts 9–12 — workbook, package lifecycle and content validation
# --------------------------------------------------------------------------- #

def test_openpyxl_is_declared_for_the_backend_and_project():
    text = (ROOT / "webapp" / "backend" / "requirements.txt").read_text()
    assert "openpyxl" in text
    project = tomllib.loads((ROOT / "pyproject.toml").read_text())
    assert any(
        dependency.lower().startswith("openpyxl")
        for dependency in project["project"]["dependencies"]
    )


def test_workbook_capability_is_reported_not_raised():
    from exondomaincompare.shared_gene_analysis.package_builder import workbook_capability
    ok, detail = workbook_capability()
    assert isinstance(ok, bool)
    if not ok:
        assert "pip install" in detail


def test_no_raw_dependency_error_reaches_the_user():
    from exondomaincompare.shared_gene_analysis.package_builder import WORKBOOK_FAILURE_MESSAGE
    assert "openpyxl" not in WORKBOOK_FAILURE_MESSAGE
    assert WORKBOOK_FAILURE_MESSAGE.startswith("Workbook generation failed.")


@pytest.mark.parametrize("scope", ["comparative", GALLUS, MUS, "all"])
def test_package_lifecycle_produces_a_valid_downloadable_zip(scope):
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import build_package, validate_package
    job = build_package(MULTI, {"scope": scope, "preset": "recommended"})
    assert job.status == "ready", job.error
    assert job.package_name.endswith(".zip")
    assert job.progress == 1.0
    zip_path = _package_path(job)
    assert zip_path.is_file() and zip_path.stat().st_size > 1000
    with zipfile.ZipFile(zip_path) as zf:
        names = zf.namelist()
        manifest = json.loads(
            zf.read(next(n for n in names if n.endswith("manifest.json"))))
        assert validate_package(zip_path, manifest) == []
        # Part 12
        assert not any(bad in n for n in names
                       for bad in ("__MACOSX", ".DS_Store"))
        assert not any(n.startswith("/") or ".." in n.split("/") for n in names)
        for info in zf.infolist():
            assert info.is_dir() or info.file_size > 0, info.filename
        assert len(names) == len(set(names))
        # The manifest describes exactly what is in the ZIP.
        declared = {e["arcname"] for e in manifest["included"]}
        control = {n for n in names if n.rsplit("/", 1)[-1] in
                   ("manifest.json", "README.md", "package_selection.json")}
        assert declared | control == set(names)
        # No personal path and no credential leaks into the package text.
        for name in names:
            if name.endswith((".json", ".md")):
                text = zf.read(name).decode("utf-8", "replace")
                assert "/Users/" not in text
                assert "lrz" not in text.lower() or "password" not in text.lower()


def test_species_scope_package_contains_only_that_species():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import build_package
    job = build_package(MULTI, {"scope": MUS, "preset": "recommended"})
    assert job.status == "ready", job.error
    with zipfile.ZipFile(_package_path(job)) as zf:
        payload = [n for n in zf.namelist()
                   if n.rsplit("/", 1)[-1] not in
                   ("manifest.json", "README.md", "package_selection.json")]
        assert payload
        assert all(f"/{MUS}/" in n for n in payload), [
            n for n in payload if f"/{MUS}/" not in n][:5]
        assert not any(GALLUS in n for n in payload)


def test_comparative_workbook_has_a_boundary_observations_sheet():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import build_package, workbook_capability
    if not workbook_capability()[0]:
        pytest.skip("workbook support not installed in this interpreter")
    job = build_package(MULTI, {"scope": "comparative", "preset": "recommended"})
    assert job.status == "ready", job.error
    assert not job.warnings, job.warnings
    with zipfile.ZipFile(_package_path(job)) as zf:
        xlsx = next(n for n in zf.namelist() if n.endswith(".xlsx"))
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx)))
        assert "Boundary observations" in wb.sheetnames
        sheet = wb["Boundary observations"]
        header = [c.value for c in next(sheet.iter_rows(max_row=1))]
        assert header == REQUIRED_BOUNDARY_FIELDS
        assert sheet.max_row > 1


def test_workbook_failure_does_not_stop_the_rest_of_the_package(monkeypatch):
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis import package_builder as pb
    monkeypatch.setattr(pb, "_write_workbook", _raise_workbook_error)
    job = pb.build_package(MULTI, {"scope": "comparative", "preset": "recommended"})
    assert job.status == "ready", job.error
    assert pb.WORKBOOK_FAILURE_MESSAGE in job.warnings
    with zipfile.ZipFile(_package_path(job)) as zf:
        names = zf.namelist()
        assert not any(n.endswith(".xlsx") for n in names)
        assert any("species_inventory.tsv" in n for n in names)


def _raise_workbook_error(*_args, **_kwargs):
    raise ImportError("No module named 'openpyxl'")


def test_expired_package_is_reported_as_expired():
    if not MULTI.is_dir():
        pytest.skip("multi-species run required")
    from exondomaincompare.shared_gene_analysis.package_builder import build_package, get_job, _JOBS
    job = build_package(MULTI, {"scope": "comparative", "preset": "figures"})
    assert job.status == "ready", job.error
    _package_path(job).unlink()
    _JOBS.pop(job.job_id, None)
    assert get_job(job.job_id).status == "expired"


# --------------------------------------------------------------------------- #
# Regression
# --------------------------------------------------------------------------- #

def test_tp53_danio_run_still_resolves_its_own_scope():
    if not (TP53 / "website_indices" / "generic"
            / "protein_coordinate_model.json").is_file():
        pytest.skip("TP53 Danio run not present")
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities
    caps = capabilities(TP53)
    assert caps["scopes"]
    assert caps["scope"] == caps["scopes"][0]["id"]
    for item in caps["items"].values():
        assert item["available"] or item["reason"]


def test_fgfr2_example_keeps_its_validated_files_view():
    page = (ROOT / "webapp" / "frontend" / "src" / "pages"
            / "DataDownloads.jsx").read_text()
    assert 'eventType === "validated"' in page
    assert "LegacyFgfr2Files" in page


def test_fgfr2_freeze_is_not_touched_by_the_correction():
    freeze = (ROOT / "results" / "final_30_until_interpro_prepare"
              / "13_final_pre_interpro_closure")
    if not freeze.is_dir():
        pytest.skip("FGFR2 freeze not present")
    before = {p.name: p.stat().st_mtime for p in freeze.rglob("*.json")}
    from exondomaincompare.shared_gene_analysis.package_builder import capabilities
    if MULTI.is_dir():
        capabilities(MULTI, "comparative")
    if SINGLE.is_dir():
        capabilities(SINGLE)
    after = {p.name: p.stat().st_mtime for p in freeze.rglob("*.json")}
    assert before == after
