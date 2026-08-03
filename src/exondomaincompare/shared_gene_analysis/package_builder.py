#!/usr/bin/env python3
from __future__ import annotations

import json
import hashlib
import logging
import re
import tempfile
import time
import zipfile
from dataclasses import dataclass, field, asdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Sequence, Set, Tuple
from urllib.parse import unquote

from exondomaincompare.config import discover_repository_root, load_config
from exondomaincompare.contracts import file_sha256, stamp_payload
from exondomaincompare.runs.legacy import LegacyRunAdapter
from exondomaincompare.runs.registry import discover_runs

ROOT = discover_repository_root(__file__)
RUNTIME_CONFIG = load_config(repository_root=ROOT)
ROOT = RUNTIME_CONFIG.repository_root
RUNS_ROOT = RUNTIME_CONFIG.runs_root
PACKAGES_ROOT = RUNTIME_CONFIG.paths.packages
LOG = logging.getLogger("package_builder")

SCOPE_COMPARATIVE = "comparative"
SCOPE_ALL = "all"

# Shown to the user when the workbook could not be produced. The technical
# exception is logged, never surfaced.
WORKBOOK_FAILURE_MESSAGE = (
    "Workbook generation failed. Other selected files can still be packaged.")

# Names that must never appear inside a delivered package (Part 12).
FORBIDDEN_ZIP_PATTERNS = ("__MACOSX", ".DS_Store", "Thumbs.db")


# --------------------------------------------------------------------------- #
# Item model
# --------------------------------------------------------------------------- #

@dataclass(frozen=True)
class Item:
    """One download item.

    ``resolve`` returns the real source files for a context, or an empty list
    together with the exact reason the item cannot be delivered.
    """
    id: str
    label: str
    group: str
    zip_path: str
    resolve: Callable[["Context"], "Resolution"]
    depends_on: Tuple[str, ...] = ()
    description: str = ""


@dataclass
class Resolution:
    paths: List[Path] = field(default_factory=list)
    reason: str = ""
    # Rows written on the fly (boundary observations, species-filtered extracts)
    # rather than copied from an existing file.
    derived: bool = False

    @property
    def available(self) -> bool:
        return bool(self.paths)

    @property
    def nbytes(self) -> int:
        return sum(p.stat().st_size for p in self.paths if p.is_file())


@dataclass
class Context:
    run_dir: Path
    scope: str
    gene: str
    models: List[Dict[str, Any]]
    comparative: Dict[str, Any]
    species_id: str = ""

    @property
    def generic(self) -> Path:
        return self.run_dir / "results" / "generic_gene_analysis"

    @property
    def core(self) -> Path:
        return self.run_dir / "results" / "core_gene_analysis"

    @property
    def multi_species(self) -> bool:
        return len(self.models) >= 2

    def model(self) -> Dict[str, Any]:
        for m in self.models:
            if m.get("species_id") == self.species_id:
                return m
        return {}


GROUP_LABELS: Dict[str, str] = {
    "overview": "Dataset overview",
    "sequences": "Sequences and models",
    "isoform_alignment": "Isoform alignment",
    "exon_structure": "Exon structure",
    "domain_architecture": "Domain architecture",
    "boundaries": "Exon–domain boundaries",
    "genomic_context": "Genomic context",
    "candidates": "Candidates",
    "figures": "Figures",
    "workbook": "Workbook",
    "qc": "QC",
}

SPECIES_GROUP_ORDER = ("overview", "sequences", "isoform_alignment", "exon_structure",
                       "domain_architecture", "boundaries", "genomic_context",
                       "candidates", "figures", "qc")
COMPARATIVE_GROUP_ORDER = ("overview", "sequences", "exon_structure",
                           "domain_architecture", "boundaries", "genomic_context",
                           "candidates", "figures", "workbook")


# --------------------------------------------------------------------------- #
# Resolver helpers
# --------------------------------------------------------------------------- #

def _load_json(path: Path) -> Any:
    try:
        return json.loads(Path(path).read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return None


def _file(path: Path, missing: str) -> Resolution:
    p = Path(path)
    if p.is_file() and p.stat().st_size > 0:
        return Resolution(paths=[p])
    return Resolution(reason=missing)


def _artefact(ctx: Context, key: str, label: str) -> Resolution:
    rel = (ctx.comparative.get("artefacts") or {}).get(key)
    if not rel:
        return Resolution(reason=f"{label} has not been generated for this run.")
    return _file(ROOT / rel, f"{label} is registered but the file is missing or empty.")


_TMP_FILES: List[Path] = []


def _tmp_tsv(name: str, lines: Sequence[str]) -> Path:
    tmp = Path(tempfile.mkdtemp(prefix="pkg_")) / name
    tmp.write_text("\n".join(lines) + "\n", encoding="utf-8")
    _TMP_FILES.append(tmp)
    return tmp


def _species_rows(src: Path, species_id: str, out_name: str,
                  label: str) -> Resolution:
    """A species-filtered copy of a run-level TSV that carries a species column."""
    if not src.is_file():
        return Resolution(reason=f"{label} has not been generated for this run.")
    lines = src.read_text(encoding="utf-8").splitlines()
    if not lines:
        return Resolution(reason=f"{label} is empty for this run.")
    cols = lines[0].split("\t")
    if "species_id" not in cols:
        # Run-level table without a species column: deliver it unfiltered.
        return Resolution(paths=[src])
    idx = cols.index("species_id")
    kept = [ln for ln in lines[1:] if ln.split("\t")[idx:idx + 1] == [species_id]]
    if not kept:
        return Resolution(reason=f"{label} contains no rows for this species.")
    if len(kept) == len(lines) - 1:
        # Every row already belongs to this species (the single-species case), so
        # the original file is delivered and stays directly downloadable.
        return Resolution(paths=[src])
    return Resolution(paths=[_tmp_tsv(out_name, [lines[0], *kept])], derived=True)


def _species_fasta(src: Path, species_id: str, out_name: str,
                   label: str) -> Resolution:
    """Records of one species from a run-level FASTA (`>ACC GENE|species_id`)."""
    if not src.is_file():
        return Resolution(reason=f"{label} has not been generated for this run.")
    all_lines = src.read_text(encoding="utf-8").splitlines()
    records: List[str] = []
    keep = False
    for line in all_lines:
        if line.startswith(">"):
            keep = species_id in line
        if keep:
            records.append(line)
    if not records:
        return Resolution(reason=f"{label} contains no records for this species.")
    if len(records) == len([ln for ln in all_lines if ln.strip()]):
        return Resolution(paths=[src])
    tmp = Path(tempfile.mkdtemp(prefix="pkg_")) / out_name
    tmp.write_text("\n".join(records) + "\n", encoding="utf-8")
    _TMP_FILES.append(tmp)
    return Resolution(paths=[tmp], derived=True)


def _card_file(run_dir: Path, url: str) -> Optional[Path]:
    """The file behind a Gallery card URL.

    The index stores serving URLs of the form
    ``/api/runs/<id>/files?path=<run-relative path>``; the package needs the file
    on disk, so the run-relative path is read back out of the query.
    """
    if not url:
        return None
    match = re.search(r"[?&]path=([^&]+)", url)
    rel = unquote(match.group(1)) if match else url
    candidate = Path(run_dir) / rel if not rel.startswith("runs/") else ROOT / rel
    return candidate if candidate.is_file() else None


def _species_figures(ctx: Context) -> Resolution:
    """Every registered Gallery figure file of one species, in every format."""
    index = _load_json(ctx.run_dir / "website_indices" / "figures_index.json") or {}
    wanted: List[Path] = []
    for card in index.get("figures") or []:
        if (card.get("scope") or "") == SCOPE_COMPARATIVE:
            continue
        if (card.get("species_id") or "") != ctx.species_id:
            continue
        for key in ("svg_url", "pdf_url", "png_url", "table_url", "caption_url"):
            p = _card_file(ctx.run_dir, card.get(key) or "")
            if p and p not in wanted:
                wanted.append(p)
    if not wanted:
        return Resolution(reason="No figures have been rendered for this species yet.")
    return Resolution(paths=wanted)


def _comparative_figures(ctx: Context) -> Resolution:
    fig_dir = ctx.generic / "figures" / "comparative"
    files = sorted(p for p in fig_dir.glob("*")
                   if p.is_file() and p.suffix.lower() in
                   {".svg", ".pdf", ".png", ".tsv", ".txt"}) if fig_dir.is_dir() else []
    if not files:
        return Resolution(reason="Comparative figures have not been rendered yet.")
    return Resolution(paths=files)


def _boundary_observations(ctx: Context) -> Resolution:
    """Part 8: built from the canonical JSON, never from a pre-existing TSV."""
    from exondomaincompare.shared_gene_analysis import boundary_observations as bo
    out = bo.ensure_table(ctx.run_dir, ctx.species_id or None)
    if out is None:
        return Resolution(reason=(
            "No exon–domain Boundary observation exists in the canonical "
            "coordinate model for this scope."))
    return Resolution(paths=[out], derived=True)


def _workbook(ctx: Context) -> Resolution:
    """Availability of the workbook, not the workbook itself; it is built lazily."""
    if not ctx.comparative.get("available"):
        return Resolution(reason="The workbook needs a multi-species comparative dataset.")
    ok, detail = workbook_capability()
    if not ok:
        return Resolution(reason=detail)
    # A marker path: the file is generated during the build, not read from disk.
    return Resolution(paths=[], reason="", derived=True)


def workbook_capability() -> Tuple[bool, str]:
    """Whether this interpreter can write the XLSX workbook.

    Called at application startup as well, so a missing declared dependency is
    reported once on the server rather than as a raw ImportError per request.
    """
    try:
        import openpyxl  # noqa: F401
    except Exception as exc:  # pragma: no cover - environment specific
        LOG.warning("workbook capability unavailable: %s", exc)
        return False, ("Workbook support is not installed in the server "
                       "environment (pip install -r webapp/backend/requirements.txt).")
    return True, ""


# --------------------------------------------------------------------------- #
# Catalogues
# --------------------------------------------------------------------------- #

def _species_catalogue() -> Dict[str, Item]:
    def it(iid: str, label: str, group: str, zip_path: str,
           resolve: Callable[[Context], Resolution],
           depends_on: Tuple[str, ...] = (), description: str = "") -> Tuple[str, Item]:
        return iid, Item(iid, label, group, zip_path, resolve, depends_on, description)

    return dict([
        it("species_summary", "Dataset summary (JSON)", "overview",
           "{sp}/dataset_summary.json",
           lambda c: _file(c.run_dir / "website_indices" / "dataset_summary.json",
                           "The dataset summary index has not been written yet."),
           description="Run-level summary of the analysed dataset."),
        it("coordinate_model", "Protein-coordinate model (JSON)", "overview",
           "{sp}/protein_coordinate_model.json",
           _model_for_species,
           description="The canonical per-species coordinate model behind every view."),
        it("primary_protein_fasta", "Primary protein (FASTA)", "sequences",
           "{sp}/sequences/primary_protein.faa",
           lambda c: _species_fasta(c.core / "proteins_primary.faa", c.species_id,
                                    "primary_protein.faa", "The primary-protein FASTA")),
        it("all_isoforms_fasta", "All protein models (FASTA)", "sequences",
           "{sp}/sequences/all_protein_models.faa",
           lambda c: _species_fasta(c.core / "proteins_all_isoforms.faa", c.species_id,
                                    "all_protein_models.faa",
                                    "The all-isoform protein FASTA")),
        it("protein_isoform_index", "Protein-model index (TSV)", "sequences",
           "{sp}/sequences/protein_model_index.tsv",
           lambda c: _species_rows(c.core / "protein_isoform_index.tsv", c.species_id,
                                   "protein_model_index.tsv", "The protein-model index")),
        it("gene_model_index", "Gene-model index (TSV)", "sequences",
           "{sp}/sequences/gene_model_index.tsv",
           lambda c: _species_rows(c.core / "gene_model_index.tsv", c.species_id,
                                   "gene_model_index.tsv", "The gene-model index")),
        it("isoform_alignment", "Within-species isoform alignment (FASTA)",
           "isoform_alignment", "{sp}/isoform_alignment/isoform_msa.aln.faa",
           lambda c: _file(c.generic / "msa" / f"isoform_msa__{c.species_id}.aln.faa",
                           "No within-species isoform alignment was produced for "
                           "this species.")),
        it("isoform_alignment_input", "Isoform alignment input (FASTA)",
           "isoform_alignment", "{sp}/isoform_alignment/isoform_msa_input.faa",
           lambda c: _file(c.generic / "msa" / f"isoform_msa_input__{c.species_id}.faa",
                           "No isoform alignment input was written for this species."),
           depends_on=("isoform_alignment",)),
        it("exon_protein_map", "Exon–protein architecture (TSV)", "exon_structure",
           "{sp}/exon_structure/exon_protein_architecture.tsv",
           lambda c: _species_rows(c.generic / "exon_protein_architecture.tsv",
                                   c.species_id, "exon_protein_architecture.tsv",
                                   "The exon–protein architecture table")),
        it("transcript_evidence", "Transcript-model evidence (TSV)", "exon_structure",
           "{sp}/exon_structure/transcript_model_evidence.tsv",
           lambda c: _species_rows(c.generic / "transcript_model_evidence.tsv",
                                   c.species_id, "transcript_model_evidence.tsv",
                                   "The transcript-model evidence table")),
        it("domain_architecture", "Representative domains (TSV)", "domain_architecture",
           "{sp}/domain_architecture/domain_architecture.tsv",
           lambda c: _species_rows(c.generic / "domain_architecture.tsv", c.species_id,
                                   "domain_architecture.tsv",
                                   "The domain-architecture table")),
        it("tm_features", "Transmembrane features (TSV)", "domain_architecture",
           "{sp}/domain_architecture/tm_features.tsv",
           lambda c: _species_rows(c.core / "tm_features.tsv", c.species_id,
                                   "tm_features.tsv", "The transmembrane-feature table")),
        it("boundary_observations", "Boundary observations (TSV)", "boundaries",
           "{sp}/boundaries/exon_domain_boundary_observations.tsv",
           _boundary_observations,
           description="One row per exon–domain Boundary observation of this species."),
        it("boundary_analysis", "Boundary analysis detail (TSV)", "boundaries",
           "{sp}/boundaries/exon_domain_boundary_analysis.tsv",
           lambda c: _species_rows(c.generic / "exon_domain_boundary_analysis.tsv",
                                   c.species_id, "exon_domain_boundary_analysis.tsv",
                                   "The Boundary analysis table")),
        it("synteny", "Local gene neighbourhood (TSV)", "genomic_context",
           "{sp}/genomic_context/synteny_neighbourhood.tsv",
           lambda c: _species_rows(c.generic / "synteny_neighbourhood.tsv", c.species_id,
                                   "synteny_neighbourhood.tsv",
                                   "The synteny neighbourhood table")),
        it("candidate_ranking", "Exploratory candidate ranking (TSV)", "candidates",
           "{sp}/candidates/event_candidate_ranking.tsv",
           lambda c: _species_rows(c.generic / "event_candidate_ranking.tsv",
                                   c.species_id, "event_candidate_ranking.tsv",
                                   "The candidate ranking table")),
        it("candidate_domain_context", "Candidate domain context (TSV)", "candidates",
           "{sp}/candidates/candidate_domain_context.tsv",
           lambda c: _species_rows(c.generic / "candidate_domain_context.tsv",
                                   c.species_id, "candidate_domain_context.tsv",
                                   "The candidate domain-context table"),
           depends_on=("candidate_ranking",)),
        it("species_figures", "Figures (SVG/PDF/PNG + source tables)", "figures",
           "{sp}/figures/", _species_figures,
           description="Every Figure Gallery card of this species, in every format."),
        it("qc_report", "Run QC summary (JSON)", "qc", "{sp}/qc/run_qc_summary.json",
           lambda c: _file(c.run_dir / "results" / "09_qc" / "run_qc_summary.json",
                           "No QC summary was written for this run.")),
        it("analysis_status", "Analysis status (JSON)", "qc",
           "{sp}/qc/analysis_status.json",
           lambda c: _file(c.run_dir / "status.json",
                           "No run status file is present.")),
    ])


def _model_for_species(ctx: Context) -> Resolution:
    model = ctx.model()
    if not model:
        return Resolution(reason="No coordinate model exists for this species.")
    tmp = Path(tempfile.mkdtemp(prefix="pkg_")) / "protein_coordinate_model.json"
    tmp.write_text(json.dumps(model, indent=2), encoding="utf-8")
    _TMP_FILES.append(tmp)
    return Resolution(paths=[tmp], derived=True)


def _comparative_catalogue() -> Dict[str, Item]:
    def it(iid: str, label: str, group: str, zip_path: str,
           resolve: Callable[[Context], Resolution],
           depends_on: Tuple[str, ...] = (), description: str = "") -> Tuple[str, Item]:
        return iid, Item(iid, label, group, zip_path, resolve, depends_on, description)

    return dict([
        it("species_inventory", "Species inventory (TSV)", "overview",
           "comparative/species_inventory.tsv",
           lambda c: _artefact(c, "species_inventory", "The species inventory")),
        it("analysis_availability", "Analysis availability (TSV)", "overview",
           "comparative/analysis_availability.tsv",
           lambda c: _artefact(c, "analysis_availability",
                               "The analysis-availability table")),
        it("primary_proteins_fasta", "Primary proteins (FASTA)", "sequences",
           "comparative/sequences/primary_proteins.faa",
           lambda c: _file(c.generic / "msa" / "primaries_msa_input.faa",
                           "The cross-species primary-protein FASTA is missing.")),
        it("primary_proteins_msa", "Cross-species primary-protein MSA (FASTA)",
           "sequences", "comparative/sequences/primary_proteins_aligned.faa",
           lambda c: _file(c.generic / "msa" / "primaries_msa.aln.faa",
                           "The cross-species MSA has not been computed."),
           depends_on=("primary_proteins_fasta",)),
        it("msa_aligned_exons", "MSA-aligned exons (TSV)", "exon_structure",
           "comparative/exon_structure/msa_aligned_exons.tsv",
           lambda c: _artefact(c, "msa_aligned_exons", "The MSA-aligned exon table"),
           depends_on=("primary_proteins_msa",)),
        it("comparable_boundary_groups", "Comparable-boundary groups (TSV)",
           "exon_structure", "comparative/exon_structure/comparable_boundary_groups.tsv",
           lambda c: _artefact(c, "comparable_boundary_groups",
                               "The comparable-boundary group table")),
        it("msa_aligned_domains", "MSA-aligned domains (TSV)", "domain_architecture",
           "comparative/domain_architecture/msa_aligned_domains.tsv",
           lambda c: _artefact(c, "msa_aligned_domains", "The MSA-aligned domain table"),
           depends_on=("primary_proteins_msa",)),
        it("domain_annotation_matrix", "Domain annotation matrix (TSV)",
           "domain_architecture",
           "comparative/domain_architecture/domain_annotation_matrix.tsv",
           lambda c: _artefact(c, "domain_annotation_matrix",
                               "The domain annotation matrix"),
           depends_on=("msa_aligned_domains",)),
        it("comparable_domain_groups", "Comparable-domain groups (TSV)",
           "domain_architecture",
           "comparative/domain_architecture/comparable_domain_groups.tsv",
           lambda c: _artefact(c, "comparable_domain_groups",
                               "The comparable-domain group table"),
           depends_on=("msa_aligned_domains",)),
        it("boundary_long_table", "All species Boundary observations (TSV)",
           "boundaries", "comparative/boundaries/exon_domain_boundaries_long.tsv",
           _boundary_observations,
           description="One row per species-specific Boundary observation, with its "
                       "comparable-boundary group and mapping confidence."),
        it("boundary_consistency", "Boundary consistency summary (TSV)", "boundaries",
           "comparative/boundaries/boundary_consistency_summary.tsv",
           lambda c: _artefact(c, "boundary_consistency_summary",
                               "The Boundary consistency summary"),
           depends_on=("boundary_long_table",)),
        it("inspection_cases", "Boundary inspection cases (TSV)", "boundaries",
           "comparative/boundaries/inspection_cases.tsv",
           lambda c: _artefact(c, "inspection_cases", "The inspection-case table"),
           depends_on=("boundary_long_table",)),
        it("comparative_synteny", "Comparative synteny (TSV)", "genomic_context",
           "comparative/synteny/comparative_synteny.tsv",
           lambda c: _artefact(c, "comparative_synteny", "The comparative synteny table")),
        it("isoform_diversity", "Isoform diversity summary (TSV)", "candidates",
           "comparative/candidates/isoform_diversity.tsv",
           lambda c: _artefact(c, "isoform_diversity", "The isoform diversity summary")),
        it("comparative_figures", "Comparative figures (SVG/PDF/PNG + source tables)",
           "figures", "comparative/figures/", _comparative_figures),
        it("excel_workbook", "Comparative Excel workbook (XLSX)", "workbook",
           "{gene}_comparative_results.xlsx", _workbook,
           depends_on=("species_inventory", "analysis_availability",
                       "msa_aligned_exons", "msa_aligned_domains",
                       "domain_annotation_matrix", "boundary_long_table",
                       "boundary_consistency", "comparative_synteny"),
           description="Every comparative table as one workbook, one sheet per table."),
    ])


SPECIES_CATALOGUE = _species_catalogue()
COMPARATIVE_CATALOGUE = _comparative_catalogue()

# Preset -> item ids. A preset is a *request*: capability filtering removes the
# items the run cannot deliver, so a preset never preselects an unavailable box.
SPECIES_PRESETS: Dict[str, Dict[str, Any]] = {
    "recommended": {
        "label": "Recommended",
        "description": "The tables, sequences and figures needed to reproduce the "
                       "species pages.",
        "items": ["species_summary", "coordinate_model", "primary_protein_fasta",
                  "all_isoforms_fasta", "protein_isoform_index", "isoform_alignment",
                  "exon_protein_map", "domain_architecture", "tm_features",
                  "boundary_observations", "boundary_analysis", "synteny",
                  "candidate_ranking", "species_figures", "qc_report"],
    },
    "tables_alignments": {
        "label": "Tables & alignments",
        "description": "Machine-readable tables and the sequence alignments, no figures.",
        "items": ["protein_isoform_index", "gene_model_index", "primary_protein_fasta",
                  "all_isoforms_fasta", "isoform_alignment", "exon_protein_map",
                  "transcript_evidence", "domain_architecture", "tm_features",
                  "boundary_observations", "boundary_analysis", "synteny",
                  "candidate_ranking", "candidate_domain_context"],
    },
    "figures": {
        "label": "Figures",
        "description": "Every Gallery figure of this species with its source tables.",
        "items": ["species_figures"],
    },
    "custom": {"label": "Custom", "description": "Pick the individual files.",
               "items": []},
}

COMPARATIVE_PRESETS: Dict[str, Dict[str, Any]] = {
    "recommended": {
        "label": "Recommended",
        "description": "The comparative tables, the cross-species alignment, the "
                       "comparative figures and the workbook.",
        "items": ["species_inventory", "analysis_availability",
                  "primary_proteins_fasta", "primary_proteins_msa",
                  "msa_aligned_exons", "comparable_boundary_groups",
                  "msa_aligned_domains", "domain_annotation_matrix",
                  "boundary_long_table", "boundary_consistency", "inspection_cases",
                  "comparative_synteny", "isoform_diversity",
                  "comparative_figures", "excel_workbook"],
    },
    "tables_alignments": {
        "label": "Tables & alignments",
        "description": "Comparative tables and the cross-species MSA, no figures.",
        "items": ["species_inventory", "analysis_availability",
                  "primary_proteins_fasta", "primary_proteins_msa",
                  "msa_aligned_exons", "comparable_boundary_groups",
                  "msa_aligned_domains", "domain_annotation_matrix",
                  "comparable_domain_groups", "boundary_long_table",
                  "boundary_consistency", "comparative_synteny", "isoform_diversity"],
    },
    "figures": {
        "label": "Figures",
        "description": "The comparative figures with their source tables.",
        "items": ["comparative_figures"],
    },
    "custom": {"label": "Custom", "description": "Pick the individual files.",
               "items": []},
}


def catalogue_for_scope(scope: str) -> Dict[str, Item]:
    return COMPARATIVE_CATALOGUE if scope == SCOPE_COMPARATIVE else SPECIES_CATALOGUE


def presets_for_scope(scope: str) -> Dict[str, Dict[str, Any]]:
    return COMPARATIVE_PRESETS if scope == SCOPE_COMPARATIVE else SPECIES_PRESETS


def resolve_dependencies(item_ids: Sequence[str], scope: str) -> List[str]:
    """Item ids with their dependencies pulled in, in a stable build order."""
    catalogue = catalogue_for_scope(scope)
    seen: Set[str] = set()
    ordered: List[str] = []

    def add(iid: str) -> None:
        if iid in seen or iid not in catalogue:
            return
        seen.add(iid)
        for dep in catalogue[iid].depends_on:
            add(dep)
        if iid not in ordered:
            ordered.append(iid)

    for iid in item_ids:
        add(iid)
    # Dependencies first: a dependency added after its dependant is moved ahead.
    ordered.sort(key=lambda i: len(catalogue[i].depends_on))
    return ordered


# --------------------------------------------------------------------------- #
# Capability contract (Part 7)
# --------------------------------------------------------------------------- #

def _context(run_dir: Path, scope: str, species_id: str = "") -> Context:
    # Absolute, so a run passed in relative form still yields project-relative
    # download paths rather than silently degrading to package-only items.
    run_dir = Path(run_dir).resolve()
    comparative = _load_json(
        run_dir / "website_indices" / "generic"
        / "comparative_dataset_index.json") or {}
    pcm = _load_json(
        run_dir / "website_indices" / "generic"
        / "protein_coordinate_model.json") or {}
    models = pcm.get("models") or []
    gene = pcm.get("gene_symbol") or comparative.get("gene_symbol") or "GENE"
    return Context(run_dir=run_dir, scope=scope, gene=gene, models=models,
                   comparative=comparative, species_id=species_id)


def species_inventory(run_dir: Path) -> List[Dict[str, str]]:
    ctx = _context(Path(run_dir), SCOPE_COMPARATIVE)
    inv = ctx.comparative.get("species_inventory") or []
    if inv:
        return [{"species_id": r.get("species_id") or "",
                 "scientific_name": r.get("scientific_name") or r.get("species_id") or "",
                 "analysis_status": r.get("analysis_status") or ""} for r in inv
                if r.get("species_id")]
    return [{"species_id": m.get("species_id") or "",
             "scientific_name": m.get("scientific_name") or m.get("species_id") or "",
             "analysis_status": m.get("status") or ""} for m in ctx.models
            if m.get("species_id")]


def scopes_for_run(run_dir: Path) -> List[Dict[str, str]]:
    """The scopes this run really has.

    A single-species run has exactly one scope — that species — so its page can
    never show a comparative package form.
    """
    inv = species_inventory(run_dir)
    if len(inv) < 2:
        return [{"id": r["species_id"], "label": r["scientific_name"], "kind": "species"}
                for r in inv]
    return (
        [{"id": SCOPE_COMPARATIVE, "label": "Comparative package", "kind": "comparative"},
         {"id": SCOPE_ALL, "label": "All species", "kind": "all"}]
        + [{"id": r["species_id"], "label": r["scientific_name"], "kind": "species"}
           for r in inv]
    )


def capabilities(run_dir: Path, scope: Optional[str] = None) -> Dict[str, Any]:
    """The canonical availability object for one scope (Part 7).

    Every item reports ``available``, its ``path`` when it has one, the exact
    ``reason`` when it has none, and its ``dependencies``. Presets are returned
    already filtered to the available items, so selecting a preset can never
    preselect something the run cannot deliver.
    """
    run_dir = Path(run_dir)
    inv = species_inventory(run_dir)
    scopes = scopes_for_run(run_dir)
    valid = {s["id"] for s in scopes}
    scope = scope or (SCOPE_COMPARATIVE if len(inv) >= 2
                      else (inv[0]["species_id"] if inv else SCOPE_COMPARATIVE))
    if scope not in valid and scopes:
        scope = scopes[0]["id"]

    # "All species" is the species catalogue applied to every species; its
    # availability is the union, and an item is available if any species has it.
    if scope == SCOPE_ALL:
        species_ids = [r["species_id"] for r in inv]
    elif scope == SCOPE_COMPARATIVE:
        species_ids = []
    else:
        species_ids = [scope]

    catalogue = catalogue_for_scope(
        SCOPE_COMPARATIVE if scope == SCOPE_COMPARATIVE else "species")
    group_order = (COMPARATIVE_GROUP_ORDER if scope == SCOPE_COMPARATIVE
                   else SPECIES_GROUP_ORDER)

    items: Dict[str, Any] = {}
    for iid, item in catalogue.items():
        if scope == SCOPE_COMPARATIVE:
            res = _safe_resolve(item, _context(run_dir, scope))
            paths, reason, nbytes = res.paths, res.reason, res.nbytes
            per_species: Dict[str, Any] = {}
        else:
            paths, reasons, nbytes, per_species = [], [], 0, {}
            for sid in species_ids:
                res = _safe_resolve(item, _context(run_dir, scope, sid))
                per_species[sid] = {"available": res.available or _is_lazy(iid),
                                    "reason": res.reason}
                if res.available:
                    paths.extend(res.paths)
                    nbytes += res.nbytes
                elif res.reason:
                    reasons.append(f"{sid}: {res.reason}")
            reason = "" if paths else "; ".join(reasons)
        available = bool(paths) or _is_lazy(iid)
        items[iid] = {
            "id": iid,
            "label": item.label,
            "group": item.group,
            "group_label": GROUP_LABELS.get(item.group, item.group),
            "description": item.description,
            "available": available,
            # A direct-download path is offered only for a single real file on
            # disk; a bundle or a file assembled during the build is package-only.
            "path": (str(paths[0].relative_to(ROOT))
                     if len(paths) == 1 and paths[0].is_relative_to(ROOT)
                     and not _is_derived_path(paths[0]) else None),
            "n_files": len(paths),
            "reason": None if available else (reason or "Not available for this scope."),
            "dependencies": list(item.depends_on),
            "estimated_bytes": nbytes,
            "per_species": per_species,
        }

    presets = {}
    for pid, preset in presets_for_scope(
            SCOPE_COMPARATIVE if scope == SCOPE_COMPARATIVE else "species").items():
        wanted = [i for i in preset["items"] if items.get(i, {}).get("available")]
        presets[pid] = {
            "id": pid, "label": preset["label"], "description": preset["description"],
            "items": wanted,
            "unavailable_items": [i for i in preset["items"]
                                  if not items.get(i, {}).get("available")],
            "estimated_bytes": sum(items[i]["estimated_bytes"] for i in wanted),
        }

    groups = [{"id": g, "label": GROUP_LABELS.get(g, g),
               "items": [i for i, v in items.items() if v["group"] == g]}
              for g in group_order
              if any(v["group"] == g for v in items.values())]

    return {
        "run_id": run_dir.name,
        "gene_symbol": _context(run_dir, scope).gene,
        "scope": scope,
        "scopes": scopes,
        "multi_species": len(inv) >= 2,
        "species": inv,
        "selected_species": species_ids,
        "groups": groups,
        "items": items,
        "presets": presets,
        "default_preset": "recommended",
        "workbook_supported": workbook_capability()[0],
    }


def _is_lazy(item_id: str) -> bool:
    """Items generated during the build rather than resolved from disk."""
    return item_id == "excel_workbook"


def _is_derived_path(path: Path) -> bool:
    return "pkg_" in path.parts[-2] if len(path.parts) >= 2 else False


def _safe_resolve(item: Item, ctx: Context) -> Resolution:
    try:
        res = item.resolve(ctx)
    except Exception as exc:  # pragma: no cover - defensive
        LOG.warning("resolving %s failed: %s", item.id, exc)
        return Resolution(reason=f"{item.label} could not be resolved for this run.")
    if item.id == "excel_workbook" and not res.reason:
        return Resolution(paths=[], derived=True)
    return res


# Kept for the existing endpoint name; the capability object is the contract.


# --------------------------------------------------------------------------- #
# Job model
# --------------------------------------------------------------------------- #

@dataclass
class PackageJob:
    job_id: str
    run_id: str
    status: str = "validating"  # validating|building|ready|failed|expired
    preset: str = "recommended"
    scope: str = SCOPE_COMPARATIVE
    selected_items: List[str] = field(default_factory=list)
    skipped_items: List[Dict[str, str]] = field(default_factory=list)
    selected_species: List[str] = field(default_factory=list)
    include_formats: List[str] = field(
        default_factory=lambda: ["tsv", "xlsx", "faa", "json", "svg", "pdf", "png"])
    estimated_bytes: int = 0
    progress: float = 0.0
    message: str = ""
    package_name: str = ""
    zip_path: str = ""
    n_files: int = 0
    warnings: List[str] = field(default_factory=list)
    manifest: Dict[str, Any] = field(default_factory=dict)
    created_at: str = ""
    updated_at: str = ""
    error: str = ""


_JOBS: Dict[str, PackageJob] = {}


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def get_job(job_id: str) -> Optional[PackageJob]:
    if job_id in _JOBS:
        return _JOBS[job_id]
    sides = list(PACKAGES_ROOT.glob(f"*/{job_id}.json"))
    records, _ = discover_runs(RUNTIME_CONFIG)
    for record in records:
        legacy = LegacyRunAdapter(record.path).old_packages()
        sides.extend(
            path for path in legacy
            if path.name == f"{job_id}.json")
    for side in sides:
        try:
            data = json.loads(side.read_text(encoding="utf-8"))
            job = PackageJob(**{k: data[k] for k in PackageJob.__dataclass_fields__
                                if k in data})
            zp = None
            if job.zip_path:
                if job.zip_path.startswith("package:"):
                    package_ref = job.zip_path[len("package:"):]
                    zp = PACKAGES_ROOT / package_ref
                else:
                    logical = Path(job.zip_path)
                    zp = (RUNS_ROOT.joinpath(*logical.parts[1:])
                          if logical.parts and logical.parts[0] == "runs"
                          else ROOT / logical)
            if job.status == "ready" and (not zp or not zp.is_file()):
                job.status = "expired"
                job.message = "This package has expired; build it again."
            _JOBS[job_id] = job
            return job
        except Exception:  # pragma: no cover - corrupt sidecar
            continue
    return None


# --------------------------------------------------------------------------- #
# Package construction
# --------------------------------------------------------------------------- #

def _safe_arcname(name: str) -> str:
    """A ZIP entry name that can never carry an absolute or escaping path."""
    cleaned = re.sub(r"^[A-Za-z]:", "", str(name)).replace("\\", "/")
    parts = [p for p in cleaned.split("/") if p not in ("", ".", "..")]
    return "/".join(parts)


def build_package(run_dir: Path, selection: Dict[str, Any],
                  job: Optional[PackageJob] = None) -> PackageJob:
    run_dir = Path(run_dir)
    run_id = run_dir.name
    job = job or PackageJob(job_id=f"pkg_{int(time.time() * 1000)}", run_id=run_id,
                            created_at=_now())
    job.status = "validating"
    job.updated_at = _now()

    caps = capabilities(run_dir, selection.get("scope"))
    scope = caps["scope"]
    job.scope = scope
    job.preset = selection.get("preset") or caps["default_preset"]

    requested = list(selection.get("items") or [])
    if not requested:
        requested = list((caps["presets"].get(job.preset) or {}).get("items") or [])
    resolved = resolve_dependencies(requested, scope)

    # Exclude unavailable items from the resolved selection.
    skipped = [{"item": i, "label": caps["items"].get(i, {}).get("label", i),
                "reason": caps["items"][i]["reason"]}
               for i in resolved if i in caps["items"]
               and not caps["items"][i]["available"]]
    resolved = [i for i in resolved if caps["items"].get(i, {}).get("available")]
    job.selected_items = resolved
    job.skipped_items = skipped

    if not resolved:
        job.status = "failed"
        job.error = "Nothing in this selection is available for this run."
        job.message = job.error
        job.updated_at = _now()
        _JOBS[job.job_id] = job
        return job

    species_ids = list(caps["selected_species"])
    if selection.get("species"):
        species_ids = [s for s in species_ids if s in selection["species"]] or species_ids
    job.selected_species = species_ids

    ctx_gene = caps["gene_symbol"]
    catalogue = catalogue_for_scope(
        SCOPE_COMPARATIVE if scope == SCOPE_COMPARATIVE else "species")

    out_root = PACKAGES_ROOT / run_id
    out_root.mkdir(parents=True, exist_ok=True)

    scope_tag = {SCOPE_COMPARATIVE: "comparative", SCOPE_ALL: "all_species"}.get(
        scope, scope)
    package_name = f"{ctx_gene}_{run_id}_{scope_tag}_{job.preset}.zip"
    zip_path = out_root / package_name
    prefix = f"{ctx_gene}_{run_id}"

    included: List[Dict[str, Any]] = []
    omitted: List[Dict[str, Any]] = [
        {"item": s["item"], "label": s["label"], "reason": s["reason"],
         "source_status": "unavailable"} for s in skipped]
    job.status = "building"
    job.message = "Collecting files"
    job.updated_at = _now()

    def add(src: Path, arcname: str) -> None:
        if not src.is_file() or src.stat().st_size == 0:
            return
        arc = _safe_arcname(arcname)
        if any(bad in arc for bad in FORBIDDEN_ZIP_PATTERNS):
            return
        if any(e["arcname"] == arc for e in included):
            return  # never duplicate a figure or a table
        zf.write(src, arc)
        try:
            source = "run:" + str(src.resolve().relative_to(run_dir.resolve()))
        except ValueError:
            source = f"generated:{src.name}"
        included.append({
            "arcname": arc,
            "source": source,
            "bytes": src.stat().st_size,
            "sha256": file_sha256(src),
            "status": "included",
        })

    try:
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            total = max(1, len(resolved))
            for i, iid in enumerate(resolved):
                item = catalogue[iid]
                job.progress = round((i + 0.5) / total, 3)
                job.message = f"Packaging {item.label}"
                job.updated_at = _now()

                if iid == "excel_workbook":
                    ok, detail = workbook_capability()
                    if not ok:
                        omitted.append({"item": iid, "label": item.label,
                                        "reason": WORKBOOK_FAILURE_MESSAGE,
                                        "source_status": "unavailable"})
                        job.warnings.append(WORKBOOK_FAILURE_MESSAGE)
                        LOG.warning("workbook skipped: %s", detail)
                        continue
                    try:
                        with tempfile.TemporaryDirectory() as tmp:
                            xlsx = Path(tmp) / f"{ctx_gene}_comparative_results.xlsx"
                            _write_workbook(xlsx, _context(run_dir, scope).comparative,
                                            ctx_gene)
                            add(xlsx, f"{prefix}/{xlsx.name}")
                    except Exception as exc:
                        LOG.exception("workbook generation failed: %s", exc)
                        omitted.append({"item": iid, "label": item.label,
                                        "reason": WORKBOOK_FAILURE_MESSAGE,
                                        "source_status": "failed"})
                        job.warnings.append(WORKBOOK_FAILURE_MESSAGE)
                    continue

                contexts = ([_context(run_dir, scope)] if scope == SCOPE_COMPARATIVE
                            else [_context(run_dir, scope, sid) for sid in species_ids])
                produced = 0
                for ctx in contexts:
                    res = _safe_resolve(item, ctx)
                    if not res.available:
                        if res.reason and scope != SCOPE_COMPARATIVE:
                            omitted.append({
                                "item": f"{iid}:{ctx.species_id}",
                                "label": f"{item.label} · {ctx.species_id}",
                                "reason": res.reason, "source_status": "unavailable"})
                        continue
                    base = item.zip_path.format(gene=ctx_gene, sp=ctx.species_id or "")
                    for p in res.paths:
                        arc = (f"{prefix}/{base}{p.name}" if base.endswith("/")
                               else f"{prefix}/{base}")
                        add(p, arc)
                        produced += 1
                if not produced and not any(o["item"].startswith(iid)
                                            for o in omitted):
                    omitted.append({"item": iid, "label": item.label,
                                    "reason": "No file was produced for this selection.",
                                    "source_status": "unavailable"})

            job.message = "Writing manifest"
            job.progress = 0.95
            selection_doc = {
                "preset": job.preset, "scope": scope,
                "requested_items": requested, "resolved_items": resolved,
                "skipped_items": skipped, "selected_species": species_ids,
            }
            manifest = stamp_payload({
                "gene_symbol": ctx_gene,
                "dataset_id": run_id,
                "run_id": run_id,
                "generated_at": _now(),
                "preset": job.preset,
                "scope": scope,
                "selected_species": species_ids,
                "requested": requested,
                "resolved": resolved,
                "included": included,
                "omitted": omitted,
                "n_included_files": len(included),
            }, payload_type="download_manifest", run_id=run_id,
               dataset_id=run_id, profile=RUNTIME_CONFIG.public_identity(),
               generator="src/exondomaincompare/shared_gene_analysis/package_builder.py")
            readme = _readme(ctx_gene, run_id, scope, job.preset, species_ids,
                             included, omitted)
            for name, text in (
                ("package_selection.json", json.dumps(selection_doc, indent=2)),
                ("manifest.json", json.dumps(manifest, indent=2)),
                ("README.md", readme),
            ):
                zf.writestr(f"{prefix}/{name}", text)

        problems = validate_package(zip_path, manifest)
        if problems:
            raise RuntimeError("; ".join(problems))

        job.status = "ready"
        job.progress = 1.0
        job.message = "Package ready"
        job.package_name = package_name
        job.zip_path = f"package:{run_id}/{package_name}"
        job.estimated_bytes = zip_path.stat().st_size
        job.n_files = len(included) + 3
        job.manifest = {"included": len(included), "omitted": omitted,
                        "resolved_items": resolved, "selected_species": species_ids}
    except Exception as exc:
        LOG.exception("package build failed")
        job.status = "failed"
        job.error = str(exc)
        job.message = "Package build failed"
        if zip_path.exists():
            try:
                zip_path.unlink()
            except OSError:
                pass
    finally:
        job.updated_at = _now()
        for tmp in _TMP_FILES:
            try:
                tmp.unlink(missing_ok=True)
            except OSError:
                pass
        _TMP_FILES.clear()

    _JOBS[job.job_id] = job
    (out_root / f"{job.job_id}.json").write_text(json.dumps(asdict(job), indent=2),
                                                 encoding="utf-8")
    return job


def _readme(gene: str, run_id: str, scope: str, preset: str,
            species_ids: Sequence[str], included: Sequence[Dict[str, Any]],
            omitted: Sequence[Dict[str, Any]]) -> str:
    lines = [
        f"# {gene} scientific package", "",
        f"- Run: `{run_id}`",
        f"- Scope: `{scope}`",
        f"- Preset: `{preset}`",
        f"- Species: {', '.join(species_ids) if species_ids else 'comparative only'}",
        f"- Files: {len(included)}",
        "",
        "## Contents", "",
    ]
    seen: Set[str] = set()
    for entry in included:
        top = entry["arcname"].split("/")[1:2]
        key = top[0] if top else entry["arcname"]
        if key in seen:
            continue
        seen.add(key)
        lines.append(f"- `{key}`")
    if omitted:
        lines += ["", "## Not included", ""]
        lines += [f"- {o.get('label') or o['item']}: {o['reason']}" for o in omitted]
    lines += [
        "", "## Reading the tables", "",
        "Missing values use one explicit vocabulary: `unavailable`, `pending`, "
        "`not detected`, `uncertain mapping`. A blank cell is never used to mean "
        "biological absence.",
        "",
        "`manifest.json` lists every file with its source and every omitted item "
        "with the exact reason it is missing.",
        "",
    ]
    return "\n".join(lines)


def validate_package(zip_path: Path, manifest: Dict[str, Any]) -> List[str]:
    """Part 12 checks. Returns the list of problems; empty means the ZIP is good."""
    problems: List[str] = []
    with zipfile.ZipFile(zip_path, "r") as zf:
        bad = zf.testzip()
        if bad:
            problems.append(f"corrupt entry {bad}")
        names = zf.namelist()
        for name in names:
            if name.startswith("/") or ".." in name.split("/") or re.match(
                    r"^[A-Za-z]:", name):
                problems.append(f"unsafe path {name}")
            if any(p in name for p in FORBIDDEN_ZIP_PATTERNS):
                problems.append(f"forbidden entry {name}")
        for info in zf.infolist():
            if not info.is_dir() and info.file_size == 0:
                problems.append(f"empty file {info.filename}")
        declared_entries = {entry["arcname"]: entry for entry in manifest.get("included") or []}
        for name, entry in declared_entries.items():
            if name not in names or not entry.get("sha256"):
                continue
            digest = hashlib.sha256()
            with zf.open(name) as handle:
                for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                    digest.update(chunk)
            if digest.hexdigest() != entry["sha256"]:
                problems.append(f"checksum mismatch {name}")
        declared = {e["arcname"] for e in manifest.get("included") or []}
        actual = {n for n in names if not n.endswith("/")}
        control = {f"{manifest['gene_symbol']}_{manifest['run_id']}/{n}"
                   for n in ("manifest.json", "README.md", "package_selection.json")}
        missing = declared - actual
        extra = actual - declared - control
        if missing:
            problems.append(f"manifest lists {len(missing)} file(s) not in the ZIP")
        if extra:
            problems.append(f"ZIP holds {len(extra)} file(s) not in the manifest")
    return problems


# --------------------------------------------------------------------------- #
# Workbook
# --------------------------------------------------------------------------- #

def _write_workbook(path: Path, comparative: Dict[str, Any], gene: str) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    ws.append(["Sheet", "Description"])
    for row in (
        ("Species", "Species inventory with analysis status"),
        ("Analysis status", "Per-species availability of scientific layers"),
        ("Primary proteins", "Selected primary protein per species"),
        ("Isoform summary", "Compact isoform diversity comparison"),
        ("MSA-aligned exons", "Coding exons projected to MSA columns"),
        ("MSA-aligned domains", "Representative domains projected to MSA columns"),
        ("Domain annotation matrix", "detected / not detected / pending / …"),
        ("Boundary observations", "One row per species-specific Boundary observation"),
        ("Boundary consistency", "Per-group consistency statistics"),
        ("Synteny", "Comparative local genomic context"),
        ("QC", "Dataset-level notes"),
    ):
        ws.append(list(row))
    ws.append([])
    ws.append(["Missing-value vocabulary",
               "unavailable | pending | not detected | uncertain mapping"])
    ws.append(["Note", "Blank cells are not used for biological absence."])

    def cell(value: Any) -> Any:
        """A spreadsheet cell. A nested structure becomes one readable line rather
        than crashing the writer, so no sheet is lost to a list-valued column."""
        if value is None:
            return ""
        if isinstance(value, (str, int, float, bool)):
            return value
        if isinstance(value, dict):
            return "; ".join(f"{k}={cell(v)}" for k, v in value.items())
        if isinstance(value, (list, tuple)):
            return " | ".join(str(cell(v)) for v in value)
        return str(value)

    def add_sheet(name: str, rows: Sequence[Dict[str, Any]]) -> None:
        sh = wb.create_sheet(name[:31])
        if not rows:
            sh.append(["status"])
            sh.append(["unavailable"])
            return
        keys: List[str] = []
        for r in rows:
            for k in r:
                if k not in keys:
                    keys.append(k)
        sh.append(keys)
        for r in rows:
            sh.append([cell(r.get(k)) for k in keys])
        sh.auto_filter.ref = sh.dimensions
        sh.freeze_panes = "A2"
        for i, key in enumerate(keys, start=1):
            sh.column_dimensions[get_column_letter(i)].width = min(
                40, max(10, len(str(key)) + 2))

    inv = comparative.get("species_inventory") or []
    add_sheet("Species", inv)
    add_sheet("Analysis status", comparative.get("analysis_availability") or [])
    add_sheet("Primary proteins", [{
        "species_id": r.get("species_id"),
        "scientific_name": r.get("scientific_name"),
        "protein_id": r.get("protein_id"),
        "transcript_id": r.get("transcript_id"),
        "protein_length": r.get("protein_length"),
    } for r in inv])
    add_sheet("Isoform summary", comparative.get("isoform_diversity") or [])
    add_sheet("MSA-aligned exons", comparative.get("msa_aligned_exons") or [])
    add_sheet("MSA-aligned domains", comparative.get("msa_aligned_domains") or [])
    add_sheet("Domain annotation matrix",
              comparative.get("domain_annotation_matrix") or [])
    add_sheet("Boundary observations", comparative.get("boundary_long") or [])
    add_sheet("Boundary consistency", comparative.get("boundary_consistency") or [])
    add_sheet("Synteny", comparative.get("synteny") or [])
    qc = wb.create_sheet("QC")
    qc.append(["key", "value"])
    qc.append(["gene_symbol", gene])
    qc.append(["n_species", comparative.get("n_species")])
    qc.append(["msa_available", (comparative.get("msa") or {}).get("available")])
    qc.append(["msa_columns", (comparative.get("msa") or {}).get("n_columns")])
    qc.append(["generated_at", _now()])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
